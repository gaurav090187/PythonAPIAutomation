import requests
import json
import openpyxl


def test_add_student():
    api_url = 'http://thetestingworldapi.com/api/studentsDetails'
    with open('TestCase/create_student.json', 'r') as f:
        data = f.read()
    json_request = json.loads(data)

    wk = openpyxl.load_workbook('TestCase/add_student.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row

    for i in range(2, rows + 1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_birth = sh.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_birth.value
        response = requests.post(api_url, json_request)
        print(response.json())
        assert response.status_code == 201
